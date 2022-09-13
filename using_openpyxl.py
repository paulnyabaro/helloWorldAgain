import openpyxl
# Converting between column letters and numbers
# from openpyxl.cell import get_column_letter, column_index_from_string

wb = openpyxl.load_workbook('example.xlsx')
# print(wb.get_sheet_names()) # Deprecated function
print(wb.sheetnames) # The new function

# get_column_letter(900)

# Getting rows and columns from the sheets
# sheet = wb.get_sheet_by_name('Sheet1') #Deprecated
sheet = wb['Sheet1']
tuple(sheet['A1':'C3'])
for rowOfCellObjects in sheet['A1':'C3']:
    for cellObj in rowOfCellObjects:
        print(cellObj.coordinate, cellObj.value)
    print('--- END OF ROW ---')

# Writing to a file
# resultFile = open('census2010.py', 'w')
# resultFile.write('allData = ' + pprint.pformat(countyData))

# Creating a new blank workbook object
# openpyxl.Workbook()

# Creating and removing sheets
# use create_sheet() and remove_sheet() functions

# using different fonts in Excel files
# from openpyxl.styles import Font, Style

# import openpyxl
# from openpyxl.styles import Font, Style
# wb = openpyxl.Workbook()
# sheet = wb.get_sheet_by_name('Sheet')
# italic24Font = Font(size=24, italic=True)
# styleObj = Style(font=italic24Font)

# Setting values of cells from Python to Excel
# sheet['A1'] = 200
# sheet['A2'] = 300
# sheet['A3'] = '=SUM(A1:A2)'

# * Might be useful
# Columns with widths of 0 or rows with heights of 0 are hidden from the user

# Merging cells
# merge_cells()
# sheet.merge_cells('A1:D3') or sheet.merge_cells('C5:D5')

# Unmerge cells
# sheet.unmerge_cells('A1:D3')
# sheet.unmerge_cells('C5:D5')

# Freezing panes means always visible no matter the scroll, good for titles
# sheet.freeze_panes = 'A2'Row 1
# sheet.freeze_panes = 'B1'Column A
# sheet.freeze_panes = 'C1'Columns A and B
# sheet.freeze_panes = 'C2'Row 1 and columns A and B

# drawing chats
# chartObj = openpyxl.charts.BarChart()